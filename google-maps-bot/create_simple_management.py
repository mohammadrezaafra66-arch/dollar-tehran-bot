# python create_simple_management.py
import pandas as pd
import os
from datetime import datetime

def create_simple_management():
    """ایجاد فایل Excel مدیریت ساده و خوانا"""
    
    output_path = 'input/management.xlsx'
    os.makedirs('input', exist_ok=True)
    
    # ========== یک شیت واحد با همه تنظیمات ==========
    data = {
        'ردیف': range(1, 14),
        'تنظیمات': [
            '🕐 ساعت شروع',
            '🕐 ساعت پایان',
            '📅 تاریخ شروع (شمسی)',
            '📅 تاریخ پایان (شمسی)',
            '📆 روزهای اجرا',
            '🔍 حداکثر سرچ در هر شب',
            '⏱️ تأخیر بین جستجوها (ثانیه)',
            '⏱️ تأخیر بین بیزینس‌ها (ثانیه)',
            '🧵 تعداد نخ همزمان',
            '🌐 مسیر فایل پروکسی',
            '🤖 API Key کپچا',
            '⏸️ وضعیت Pause/Resume',
            '📊 حداکثر بیزینس در هر جستجو'
        ],
        'مقدار': [
            '08:00',
            '23:00',
            '1403-01-01',
            '1403-12-29',
            'شنبه,یکشنبه,دوشنبه,سه‌شنبه,چهارشنبه',
            '50',
            '30-60',
            '5-10',
            '2',
            'input/proxies.txt',
            '',
            'resume',
            '50'
        ],
        'توضیحات': [
            'ساعت شروع کار ربات (فقط بین این ساعت اجرا می‌شود)',
            'ساعت پایان کار ربات',
            'تاریخ شروع به فرمت 1403-01-01 (اختیاری)',
            'تاریخ پایان به فرمت 1403-12-29 (اختیاری)',
            'روزهایی که ربات کار کند - با کاما جدا کنید',
            'حداکثر تعداد جستجو در هر شب',
            'مثال: 30-60 یعنی بین 30 تا 60 ثانیه',
            'مثال: 5-10 یعنی بین 5 تا 10 ثانیه',
            'تعداد پردازش همزمان (پیشنهاد: 1 یا 2)',
            'مسیر فایل حاوی لیست پروکسی‌ها (یک خط یک پروکسی)',
            'API Key برای سرویس حل کپچا (در صورت نیاز)',
            'pause=توقف، resume=ادامه - قابل تغییر بدون ریستارت',
            'حداکثر بیزینس جمع‌آوری شده در هر جستجو'
        ]
    }
    
    df = pd.DataFrame(data)
    
    # ذخیره در Excel با تنظیمات ظاهری مناسب
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Config', index=False)
    
    # تنظیم ظاهر Excel
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    
    wb = load_workbook(output_path)
    ws = wb['Config']
    
    # تنظیم عرض ستون‌ها
    ws.column_dimensions['A'].width = 6   # ردیف
    ws.column_dimensions['B'].width = 35  # تنظیمات
    ws.column_dimensions['C'].width = 25  # مقدار
    ws.column_dimensions['D'].width = 55  # توضیحات
    
    # تنظیم ارتفاع ردیف‌ها
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 25
    
    # استایل برای هدر
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # استایل برای سلول‌های معمولی
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
    
    # رنگ‌بندی ردیف‌ها (یک در میان)
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        if i % 2 == 0:
            fill = PatternFill(start_color="E9F0F9", end_color="E9F0F9", fill_type="solid")
            for cell in row:
                cell.fill = fill
    
    wb.save(output_path)
    
    print("=" * 60)
    print("✅ فایل مدیریت با موفقیت ساخته شد!")
    print("=" * 60)
    print(f"📁 مسیر: {output_path}")
    print(f"📄 شیت: Management")
    print("\n📋 تنظیمات موجود:")
    print("   ✅ ساعت شروع/پایان")
    print("   ✅ تاریخ شمسی از و تا")
    print("   ✅ روزهای اجرا")
    print("   ✅ حداکثر سرچ در هر شب")
    print("   ✅ تأخیرها (بین جستجوها و بین بیزینس‌ها)")
    print("   ✅ تعداد نخ همزمان")
    print("   ✅ مسیر پروکسی")
    print("   ✅ API Key کپچا")
    print("   ✅ وضعیت Pause/Resume")
    print("   ✅ حداکثر بیزینس در هر جستجو")
    print("=" * 60)

if __name__ == "__main__":
    create_simple_management()