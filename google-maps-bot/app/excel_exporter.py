# app/excel_exporter.py - نسخه اصلاح شده (از دیتابیس می‌خواند)
import pandas as pd
import sqlite3
from datetime import datetime
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from app.config import Config

def export_to_excel(output_path='output/final_output.xlsx'):
    """
    خروجی گرفتن از دیتابیس در Excel
    (دیگر به JSON وابسته نیست)
    """
    print("=" * 60)
    print("📊 Exporting from Database to Excel")
    print("=" * 60)
    
    # اتصال مستقیم به دیتابیس
    conn = sqlite3.connect(Config.DATABASE_PATH)
    
    # گرفتن همه بیزینس‌های موفق
    df = pd.read_sql_query('''
        SELECT 
            name,
            phone,
            website,
            address,
            city,
            province,
            rating,
            reviews_count,
            place_id,
            extracted_at
        FROM businesses 
        WHERE status = 'done'
        ORDER BY id
    ''', conn)
    
    conn.close()
    
    if df.empty:
        print("⚠️ No businesses found in database!")
        # ایجاد دیتافریم خالی با ستون‌های مورد نیاز
        df = pd.DataFrame(columns=['name', 'phone', 'website', 'address', 'city', 'province'])
    
    # اضافه کردن ستون کیفیت داده
    df['data_quality'] = df.apply(lambda x: 
        'high' if x.get('phone') and x.get('website') else
        'medium' if x.get('phone') or x.get('website') else 'low', axis=1)
    
    # مرتب‌سازی بر اساس کیفیت
    df = df.sort_values('data_quality', ascending=False)
    
    # تنظیم عرض ستون‌ها
    column_widths = {
        'name': 50.00,
        'phone': 18.00,
        'website': 40.00,
        'address': 50.00,
        'city': 20.00,
        'province': 20.00,
        'rating': 10.00,
        'reviews_count': 15.00,
        'place_id': 35.00,
        'extracted_at': 22.00,
        'data_quality': 12.00
    }
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Businesses', index=False)
        
        worksheet = writer.sheets['Businesses']
        
        for idx, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(idx)
            width = column_widths.get(col, 15.00)
            worksheet.column_dimensions[col_letter].width = width
            
            for cell in worksheet[col_letter]:
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # شیت خلاصه آمار
        stats = pd.DataFrame([{
            'total_businesses': len(df),
            'has_phone': df['phone'].astype(bool).sum() if 'phone' in df.columns else 0,
            'has_website': df['website'].astype(bool).sum() if 'website' in df.columns else 0,
            'high_quality': (df['data_quality'] == 'high').sum(),
            'medium_quality': (df['data_quality'] == 'medium').sum(),
            'low_quality': (df['data_quality'] == 'low').sum(),
            'export_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }])
        stats.to_excel(writer, sheet_name='Summary', index=False)
        
        summary_ws = writer.sheets['Summary']
        for cell in summary_ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    print(f"✅ Excel saved: {output_path}")
    print(f"📊 Total businesses: {len(df)}")
    print(f"📞 Phone: {stats['has_phone'].iloc[0]}")
    print(f"🌐 Website: {stats['has_website'].iloc[0]}")

if __name__ == "__main__":
    os.makedirs('output', exist_ok=True)
    export_to_excel()