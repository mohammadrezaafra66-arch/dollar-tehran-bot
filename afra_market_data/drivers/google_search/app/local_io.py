# app/local_io.py
"""Local input/output bootstrap for Google Maps Bot.

This module keeps all user-facing inputs, settings, outputs, logs and database
inside the google-maps-bot project folder. The user edits local_settings.xlsx;
the bot converts it to the existing input Excel files before execution.
"""

from __future__ import annotations

import os
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.config import Config


LOCAL_SETTINGS_FILE = os.path.join(Config.BASE_DIR, "local_settings.xlsx")
LOCAL_README_FILE = os.path.join(Config.BASE_DIR, "LOCAL_USAGE.md")

DEFAULT_SEARCH_ROWS: List[Dict[str, object]] = [
    {
        "active": 1,
        "province": "تهران",
        "city": "تهران",
        "keyword": "فروشگاه لوازم خانگی",
        "brand": "",
        "related_keywords": "یخچال, لباسشویی, ظرفشویی",
        "category": "لوازم خانگی",
    }
]

DEFAULT_CONFIG_ROWS: List[Dict[str, object]] = [
    {"Setting": "pause_resume", "Value": "resume", "Description": "pause یا resume"},
    {"Setting": "start_time", "Value": "00:00", "Description": "ساعت شروع اجرای مجاز"},
    {"Setting": "end_time", "Value": "23:59", "Description": "ساعت پایان اجرای مجاز"},
    {"Setting": "max_scrolls", "Value": 1, "Description": "برای تست اول کم باشد"},
    {"Setting": "max_businesses_per_query", "Value": 2, "Description": "حداکثر بیزینس جمع‌آوری‌شده برای هر کوئری"},
    {"Setting": "max_businesses_to_extract", "Value": 2, "Description": "حداکثر بیزینس برای استخراج جزئیات"},
    {"Setting": "max_websites_to_crawl", "Value": 1, "Description": "حداکثر سایت برای کرال"},
    {"Setting": "headless", "Value": "FALSE", "Description": "برای تست FALSE بماند"},
    {"Setting": "slow_mo", "Value": 500, "Description": "کند کردن حرکات مرورگر بر حسب میلی‌ثانیه"},
    {"Setting": "website_crawl_enabled", "Value": "FALSE", "Description": "برای تست اول FALSE بماند"},
    {"Setting": "extract_emails", "Value": "TRUE", "Description": "استخراج ایمیل از سایت"},
    {"Setting": "extract_social", "Value": "TRUE", "Description": "استخراج شبکه‌های اجتماعی"},
    {"Setting": "delay_between_businesses", "Value": "5-10", "Description": "تأخیر بین بیزینس‌ها"},
    {"Setting": "delay_between_queries", "Value": "30-60", "Description": "تأخیر بین کوئری‌ها"},
]

DEFAULT_PHASE_ROWS: List[Dict[str, object]] = [
    {"Phase": "Phase 1", "Name": "Query Generator", "Enabled": True, "Description": "تولید کوئری از Search_Input"},
    {"Phase": "Phase 2", "Name": "Maps Collector", "Enabled": True, "Description": "جمع‌آوری نتایج گوگل مپ"},
    {"Phase": "Phase 3", "Name": "Business Extractor", "Enabled": True, "Description": "استخراج جزئیات"},
    {"Phase": "Phase 4", "Name": "Website Crawler", "Enabled": False, "Description": "کرال سایت‌ها"},
]


def _style_worksheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, column in enumerate(ws.columns, 1):
        max_len = 12
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value) + 2, 45))
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len
    ws.freeze_panes = "A2"


def create_local_settings_if_missing() -> None:
    if os.path.exists(LOCAL_SETTINGS_FILE):
        return

    os.makedirs(Config.BASE_DIR, exist_ok=True)
    with pd.ExcelWriter(LOCAL_SETTINGS_FILE, engine="openpyxl") as writer:
        pd.DataFrame(DEFAULT_SEARCH_ROWS).to_excel(writer, sheet_name="Search_Input", index=False)
        pd.DataFrame(DEFAULT_CONFIG_ROWS).to_excel(writer, sheet_name="Run_Config", index=False)
        pd.DataFrame(DEFAULT_PHASE_ROWS).to_excel(writer, sheet_name="Phases", index=False)
        pd.DataFrame([
            {"Folder": "input", "Purpose": "فایل‌های ورودی تولیدشده برای ربات"},
            {"Folder": "output", "Purpose": "خروجی Excel و JSON نهایی"},
            {"Folder": "data", "Purpose": "SQLite و checkpoint"},
            {"Folder": "logs", "Purpose": "لاگ اجرا و خطا"},
            {"Folder": "screenshots", "Purpose": "اسکرین‌شات خطا و کپچا"},
        ]).to_excel(writer, sheet_name="Local_Paths", index=False)

    wb = load_workbook(LOCAL_SETTINGS_FILE)
    for ws in wb.worksheets:
        _style_worksheet(ws)
    wb.save(LOCAL_SETTINGS_FILE)


def write_local_usage_file() -> None:
    if os.path.exists(LOCAL_README_FILE):
        return
    content = """# Google Maps Bot - Local Usage\n\nهمه چیز داخل همین پوشه انجام می‌شود.\n\n## ورودی کاربر\nفایل زیر را ویرایش کن:\n\n`local_settings.xlsx`\n\nشیت‌ها:\n- `Search_Input`: شهر، استان، کلمه اصلی، برند، کلمات مرتبط، وضعیت فعال/غیرفعال\n- `Run_Config`: تنظیمات اجرا مثل pause/resume، ساعت شروع/پایان، تعداد نتایج و تأخیرها\n- `Phases`: فعال/غیرفعال کردن فازها\n\n## خروجی کاربر\nخروجی‌ها داخل این پوشه‌ها ساخته می‌شوند:\n\n- `output/` برای فایل‌های Excel و JSON خروجی\n- `data/` برای دیتابیس و checkpoint\n- `logs/` برای لاگ‌ها\n- `screenshots/` برای عکس خطا یا کپچا\n\n## اجرا\n```powershell\npython run.py\n```\n\nبرای اجرای اول مقدارها را سبک نگه دار:\n- max_scrolls = 1\n- max_businesses_per_query = 2\n- max_businesses_to_extract = 2\n- website_crawl_enabled = FALSE\n"""
    with open(LOCAL_README_FILE, "w", encoding="utf-8") as f:
        f.write(content)


def _normalize_search_input(search_df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {
        "استان": "province",
        "شهر": "city",
        "کلمه اصلی": "keyword",
        "برند": "brand",
        "کلمات مرتبط": "related_keywords",
        "دسته بندی": "category",
        "دسته‌بندی": "category",
        "فعال": "active",
        "وضعیت": "active",
    }
    search_df = search_df.rename(columns=rename_map).copy()

    required_cols = ["active", "province", "city", "keyword", "brand", "related_keywords", "category"]
    for col in required_cols:
        if col not in search_df.columns:
            search_df[col] = "" if col != "active" else 1
    search_df = search_df[required_cols]

    search_df["active"] = pd.to_numeric(search_df["active"], errors="coerce").fillna(1).astype(int)
    for col in ["province", "city", "keyword", "brand", "related_keywords", "category"]:
        search_df[col] = search_df[col].fillna("").astype(str).replace("nan", "").str.strip()

    return search_df


def _to_legacy_input_columns(search_df: pd.DataFrame) -> pd.DataFrame:
    """Create the Persian-column Excel expected by the current QueryGenerator."""
    return pd.DataFrame({
        "active": search_df["active"],
        "استان": search_df["province"],
        "شهر": search_df["city"],
        "کلمه اصلی": search_df["keyword"],
        "برند": search_df["brand"],
        "کلمات مرتبط": search_df["related_keywords"],
        "دسته بندی": search_df["category"],
    })


def sync_local_settings_to_bot_inputs() -> None:
    """Read local_settings.xlsx and write existing bot input/management files."""
    create_local_settings_if_missing()
    write_local_usage_file()
    Config.create_directories()

    sheets = pd.read_excel(LOCAL_SETTINGS_FILE, sheet_name=None)
    search_df = sheets.get("Search_Input", pd.DataFrame(DEFAULT_SEARCH_ROWS))
    config_df = sheets.get("Run_Config", pd.DataFrame(DEFAULT_CONFIG_ROWS))
    phases_df = sheets.get("Phases", pd.DataFrame(DEFAULT_PHASE_ROWS))

    search_df = _normalize_search_input(search_df)
    legacy_input_df = _to_legacy_input_columns(search_df)

    if "Setting" in config_df.columns and "Value" in config_df.columns:
        settings = dict(zip(config_df["Setting"].astype(str), config_df["Value"]))
        needed_defaults = {row["Setting"]: row for row in DEFAULT_CONFIG_ROWS}
        for key, row in needed_defaults.items():
            if key not in settings:
                config_df = pd.concat([config_df, pd.DataFrame([row])], ignore_index=True)

    os.makedirs(Config.INPUT_DIR, exist_ok=True)
    with pd.ExcelWriter(Config.QUERIES_FILE, engine="openpyxl") as writer:
        legacy_input_df.to_excel(writer, sheet_name="Sheet1", index=False)
        search_df.to_excel(writer, sheet_name="Local_Normalized", index=False)

    with pd.ExcelWriter(Config.MANAGEMENT_FILE, engine="openpyxl") as writer:
        config_df.to_excel(writer, sheet_name="Config", index=False)
        phases_df.to_excel(writer, sheet_name="Phases", index=False)

    Config.load_from_excel()
    print(f"✅ Local settings synced: {LOCAL_SETTINGS_FILE}")
    print(f"📥 Input file: {Config.QUERIES_FILE}")
    print(f"⚙️ Management file: {Config.MANAGEMENT_FILE}")
    print(f"📤 Output folder: {Config.OUTPUT_DIR}")
