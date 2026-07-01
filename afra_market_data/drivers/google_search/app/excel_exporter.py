# app/excel_exporter.py - Google Search Driver
import pandas as pd
import sqlite3
from datetime import datetime
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from app.config import Config


def export_to_excel(output_path='output/final_output.xlsx'):
    print("=" * 60)
    print("📊 Exporting from Database to Excel")
    print("=" * 60)

    conn = sqlite3.connect(Config.DATABASE_PATH)

    df = pd.read_sql_query('''
        SELECT
            b.name,
            b.phone,
            b.website,
            b.result_url,
            b.address,
            b.city,
            b.province,
            b.result_snippet,
            b.extracted_at,
            w.email,
            w.instagram,
            w.telegram,
            w.whatsapp,
            w.extra_phones
        FROM businesses b
        LEFT JOIN website_extractions w ON w.business_id = b.id
        WHERE b.status = "done"
        ORDER BY b.id
    ''', conn)

    conn.close()

    if df.empty:
        print("⚠️ No businesses found in database!")
        df = pd.DataFrame(columns=[
            'name','phone','website','result_url','address',
            'city','province','email','result_snippet','extracted_at'
        ])

    # کیفیت داده
    df['data_quality'] = df.apply(lambda x:
        'high'   if (x.get('phone') and x.get('email')) else
        'medium' if (x.get('phone') or x.get('website')) else
        'low', axis=1)

    df = df.sort_values('data_quality', ascending=False)

    column_widths = {
        'name':           50.00,
        'phone':          18.00,
        'website':        40.00,
        'result_url':     45.00,
        'address':        50.00,
        'city':           20.00,
        'province':       20.00,
        'result_snippet': 60.00,
        'extracted_at':   22.00,
        'email':          30.00,
        'instagram':      25.00,
        'telegram':       25.00,
        'whatsapp':       18.00,
        'extra_phones':   30.00,
        'data_quality':   12.00,
    }

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else 'output', exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Businesses', index=False)

        ws = writer.sheets['Businesses']
        for idx, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = column_widths.get(col, 15.00)
            for cell in ws[col_letter]:
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

        stats = pd.DataFrame([{
            'total_businesses': len(df),
            'has_phone':    int(df['phone'].astype(bool).sum())    if 'phone'   in df.columns else 0,
            'has_email':    int(df['email'].astype(bool).sum())    if 'email'   in df.columns else 0,
            'has_website':  int(df['website'].astype(bool).sum())  if 'website' in df.columns else 0,
            'high_quality':   int((df['data_quality'] == 'high').sum()),
            'medium_quality': int((df['data_quality'] == 'medium').sum()),
            'low_quality':    int((df['data_quality'] == 'low').sum()),
            'export_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }])
        stats.to_excel(writer, sheet_name='Summary', index=False)

    print(f"✅ Excel saved: {output_path}")
    print(f"📊 Total: {len(df)}")


if __name__ == "__main__":
    os.makedirs('output', exist_ok=True)
    export_to_excel()
